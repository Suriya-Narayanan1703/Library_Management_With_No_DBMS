#Pre-processing with packages
import openpyxl
from datetime import date,datetime
today = date.today()
day=int(today.strftime("%d"))
mon=int(today.strftime("%m"))
year=int(today.strftime("%Y"))
t=datetime(year,mon,day)
xr = t.date()
print(xr)
daye=day+10
mone=mon

if(mon==1 or mon==3 or mon==5 or mon==7 or mon==8 or mon==10 or mon==12):
 if(daye>31):
    r=30-day
    d=10-r
    daye=d
    mone+=1
    if(mone>12):
        mone-=mon
        year+=1
elif(mon==2):
    if(daye>28):                                                                         
       r=28-day
       d=10-r
       daye=d
       mone+=1
       if(mone>12):
           mone-=mon
           year+=1
else:
    if(daye>30):                                                                         
       r=30-day
       d=10-r
       daye=d
       mone+=1
       if(mone>12):
           mone-=mon
           year+=1
enddate=datetime(year,mone,daye)
end=enddate.date()
#print(end)
#d2=datetime.datetime(end,"%Y/%m/%d")
#d1=datetime.datetime(today,"%Y/%m/%d")
diff=end-today
d=int(diff.days)
#print(d)

path1='Book1.xlsx'
path2='Book2.xlsx'
w=openpyxl.load_workbook(path1)
q=openpyxl.load_workbook(path2)
sheet1=w.active
sheet2=q.active
row1=sheet1.max_row
col1=sheet1.max_column
row2=sheet2.max_row
col2=sheet2.max_column
###
d1_in=sheet1.cell(row=2,column=11).value
print(d1_in.date())
d2_in=sheet1.cell(row=2,column=13).value
print(d2_in)
#print(j1)
j2=d2_in.date()
print(j2)
##modules work
#finding book
def findbook(name):
    val=0
    for i in range(1,row1+1):
       cel_sea=sheet1.cell(row=i, column=3)
       cel_sea1=sheet1.cell(row=i, column=9)
       if(cel_sea.value==name and cel_sea1.value==1):
           val=0
           cell1=sheet1.cell(row=i,column=2).value
           cell2=sheet1.cell(row=i,column=3).value
           cell3=sheet1.cell(row=i,column=5).value
           cell4=sheet1.cell(row=i,column=6).value
           print("Id:",cell1," Name:",cell2," floor:",cell3," rack:",cell4)
           break
       elif(cel_sea.value==name):
           val=1
       else:
           val=-1
           pass
    if(val==1):
        print("book is not available")
    elif(val==-1):
        print("There is no such book..")
#finding books
def findcat(name):
    for i in range(1,row1+1):
        cel=sheet1.cell(row=i, column=4)
        if(cel.value==name):
           #for j in range(1,col+1):
               cell1=sheet1.cell(row=i,column=2).value
               cell2=sheet1.cell(row=i,column=3).value
               cell3=sheet1.cell(row=i,column=5).value
               cell4=sheet1.cell(row=i,column=6).value
               print("Id:",cell1," Name:",cell2," floor:",cell3," rack:",cell4,"\n")
        else:
            pass
#MAIN WORK

while(True):
 print("\t\t\tWelcome to VIT library\t\t\t")
 print("\tA place to learn and a chance to grow\t")
 print("Name:")

 name = input()  # "Rebecca Wilcox"=input()
 print("Id:")

 id_s=input()#"20MID1"#=input()
 val=0
 mot=0
 for i in range(2,row2+1):
    cel_id=sheet2.cell(row=i, column=2)
    cel_name=sheet2.cell(row=i, column=3)
    if(cel_id.value==id_s and cel_name.value==name):
        #print("your valid")
        val=1
    else:
        pass
 if (val==1):
    print("Welcome ",name)
    print("\n1.find a book\n2.List a book of particular topic")
    print("3.Return a book\n4.Borrow a book\n5.Student Details\n6.Exit?")
    z=int(input())
    if(z==1):
        name_book=input()
        findbook(name_book)
    elif(z==2):
        name_cate=input("genre:")
        print("Listing all ",name_cate," genre books.")
        findcat(name_cate)
    elif(z==3):
          nameofbk=input()#"Data Smart"#input()
          t = datetime.today()
          for i in range(1,row1+1):
              cel=sheet1.cell(row=i, column=3)
              if(cel.value==nameofbk):
                  sheet1.cell(row=i,column=8).value=0
                  sheet1.cell(row=i,column=9).value=1
                  sheet1.cell(row=i,column=10).value=0
                  sheet1.cell(row=i,column=12).value="null"
                  sheet1.cell(row=i,column=13).value=xr
                  d1_in=sheet1.cell(row=i,column=11).value
                  print(d1_in)
                  d2_in=sheet1.cell(row=i,column=13).value
                  print(d2_in)
                  #print(j1)
                  j2=d2_in.date()
                  print(j2)
                  diff=d2_in-d1_in
                  d=int(diff.days)
                  #print(d)
                  sheet1.cell(row=i,column=11).value=0
                  sheet1.cell(row=i,column=13).value=0
                  if(d>1):
                      print("It's been",d," days you haven't returned the book")
                      print("So your are penalised...")
                      mot=(d)*10
                      #print(mot)
                  w.save('Book1.xlsx')
                  for i in range(1,row2+1):
                     cel_sea=sheet2.cell(row=i, column=3)
                     if(cel_sea.value==name):
                         sheet2.cell(row=i,column=4).value-=mot
                         #ko=sheet2.cell(row=i,column=4).value-dup
                         #sheet2.cell(row=i,column=4).value=ko
                         sheet2.cell(row=i,column=5).value="null"
                         sheet2.cell(row=i,column=6).value=0
                         q.save('Book2.xlsx')
                     else:
                         pass
              else:
                  pass
    elif(z==4):
          nameofbk=input()#"Data Smart"#input()
          for i in range(1,row1+1):
             cel_sea=sheet1.cell(row=i, column=3)
             cel_sem=sheet1.cell(row=i, column=9).value
             if(cel_sem==0):
                 print("Book not Available")
                 break
             if(cel_sea.value==nameofbk):
                 sheet1.cell(row=i,column=8).value=1
                 sheet1.cell(row=i,column=9).value=0
                 sheet1.cell(row=i,column=10).value=today
                 sheet1.cell(row=i,column=11).value=end
                 sheet1.cell(row=i,column=12).value=name
                 dup=sheet1.cell(row=i,column=8).value
                 w.save('Book1.xlsx')
                 for i in range(1,row2+1):
                    cel_set=sheet2.cell(row=i, column=3)
                    if(cel_set.value==name):
                        #ko=sheet2.cell(row=i,column=4).value-dup
                        #sheet2.cell(row=i,column=4).value=ko
                        sheet2.cell(row=i,column=5).value=nameofbk
                        sheet2.cell(row=i,column=6).value=today
                        sheet2.cell(row=i,column=7).value=end
                        q.save('Book2.xlsx')
                    else:
                        pass
             else:
                 pass
    elif(z==5):
        for i in range(1,row1+1):
           cel_sea=sheet2.cell(row=i, column=3)
           if(cel_sea.value==name):
               cell1=sheet2.cell(row=i,column=2).value
               cell2=sheet2.cell(row=i,column=3).value
               cell3=sheet2.cell(row=i,column=5).value
               print("Id:",cell1," Name:",cell2," book-bought:",cell3)
    else:
        break
 else:
    print("You are not a VIT student")



